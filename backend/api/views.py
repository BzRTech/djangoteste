from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
from django.db.models import Count, Avg, Max, Min, Q

from students.models import *
from .serializers import *

# ============================================
# VIEWSETS DE LOCALIZAÇÃO E ESTRUTURA
# ============================================

class TbCityViewSet(viewsets.ModelViewSet):
    """Cidades"""
    queryset = TbCity.objects.all()
    serializer_class = TbCitySerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['city', 'state']
    ordering_fields = ['city', 'state']


class TbSchoolViewSet(viewsets.ModelViewSet):
    """Escolas"""
    queryset = TbSchool.objects.all().select_related('id_city')
    serializer_class = TbSchoolSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['id_city', 'id_city__state']
    search_fields = ['school', 'director_name', 'address', 'codigo_ideb']


class TbTeacherViewSet(viewsets.ModelViewSet):
    """Professores com suporte a disciplinas"""
    queryset = TbTeacher.objects.all()
    serializer_class = TbTeacherSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status']
    search_fields = ['teacher_name', 'teacher_serial']
    
    @action(detail=True, methods=['get'])
    def subjects(self, request, pk=None):
        """Lista todas as disciplinas de um professor"""
        teacher = self.get_object()
        teacher_subjects = TbTeacherSubject.objects.filter(id_teacher=teacher).select_related('id_subject')
        serializer = TbTeacherSubjectSerializer(teacher_subjects, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_subject(self, request, pk=None):
        """Adiciona uma disciplina ao professor"""
        teacher = self.get_object()
        subject_id = request.data.get('subject_id')
        
        if not subject_id:
            return Response({'error': 'subject_id é obrigatório'}, status=400)
        
        try:
            subject = TbSubject.objects.get(id=subject_id)
            TbTeacherSubject.objects.get_or_create(
                id_teacher=teacher,
                id_subject=subject
            )
            return Response({'message': 'Disciplina adicionada com sucesso'})
        except TbSubject.DoesNotExist:
            return Response({'error': 'Disciplina não encontrada'}, status=404)
    
    @action(detail=True, methods=['delete'])
    def remove_subject(self, request, pk=None):
        """Remove uma disciplina do professor"""
        teacher = self.get_object()
        subject_id = request.data.get('subject_id')
        
        if not subject_id:
            return Response({'error': 'subject_id é obrigatório'}, status=400)
        
        deleted = TbTeacherSubject.objects.filter(
            id_teacher=teacher,
            id_subject_id=subject_id
        ).delete()
        
        if deleted[0] > 0:
            return Response({'message': 'Disciplina removida com sucesso'})
        return Response({'error': 'Relacionamento não encontrado'}, status=404)


class TbTeacherSubjectViewSet(viewsets.ModelViewSet):
    """Relação Professor-Disciplina"""
    queryset = TbTeacherSubject.objects.all().select_related('id_teacher', 'id_subject')
    serializer_class = TbTeacherSubjectSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_teacher', 'id_subject']


class TbTeacherSchoolViewSet(viewsets.ModelViewSet):
    """Relação Professor-Escola"""
    queryset = TbTeacherSchool.objects.all().select_related('id_teacher', 'id_school')
    serializer_class = TbTeacherSchoolSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_teacher', 'id_school', 'status']


class TbClassViewSet(viewsets.ModelViewSet):
    """Turmas"""
    queryset = TbClass.objects.all().select_related('id_teacher', 'id_school')
    serializer_class = TbClassSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['id_school', 'id_teacher', 'school_year', 'grade', 'shift']
    search_fields = ['class_name']



    def destroy(self, request, *args, **kwargs):
        """
        Sobrescreve o método destroy para verificar dependências antes de deletar
        """
        turma = self.get_object()
 
        # Verifica dependências
        dependencies = []
 
        # Verifica alunos vinculados
        students_count = TbStudents.objects.filter(id_class=turma).count()
        if students_count > 0:
            dependencies.append(f"{students_count} aluno(s)")
 
        # Verifica indicadores IDEB vinculados
        ideb_indicators_count = TbClassIdebIndicators.objects.filter(id_class=turma).count()
        if ideb_indicators_count > 0:
            dependencies.append(f"{ideb_indicators_count} indicador(es) IDEB")
 
        # Verifica aplicações de exame vinculadas
        exam_applications_count = TbExamApplications.objects.filter(id_class=turma).count()
        if exam_applications_count > 0:
            dependencies.append(f"{exam_applications_count} aplicação(ões) de exame")
 
        # Se houver dependências, retorna erro
        if dependencies:
            return Response({
                'error': 'Não é possível deletar esta turma',
                'message': f'A turma "{turma.class_name}" possui registros vinculados',
                'dependencies': dependencies,
                'suggestion': 'Remova ou reatribua os registros vinculados antes de deletar a turma'
            }, status=status.HTTP_400_BAD_REQUEST)
 
        # Se não houver dependências, permite a deleção
        return super().destroy(request, *args, **kwargs)

class TbStudentsViewSet(viewsets.ModelViewSet):
    """Alunos"""
    queryset = TbStudents.objects.all().select_related(
        'id_class',
        'id_class__id_school',
        'id_class__id_teacher'
    )
    serializer_class = TbStudentsSerializer
    lookup_field = 'id_student'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['id_class', 'status']
    search_fields = ['student_name', 'student_serial']

    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """
        Importa múltiplos estudantes via upload de arquivo CSV/Excel

        Formato esperado do arquivo (aceita português ou inglês):
        - Nome do Aluno / student_name: Nome do aluno (obrigatório)
        - Matrícula / student_serial: Matrícula (obrigatório)
        - Turma / id_class: Nome ou ID da turma (obrigatório)
        - Data de Matrícula / enrollment_date: Data de matrícula (opcional, formato YYYY-MM-DD)
        - Status / status: Status (opcional, padrão: enrolled)
        """
        import csv
        import io
        from datetime import datetime
        import re

        # Mapeamento de nomes de colunas (português -> inglês)
        COLUMN_MAPPING = {
            'nome do aluno': 'student_name',
            'nome': 'student_name',
            'aluno': 'student_name',
            'matrícula': 'student_serial',
            'matricula': 'student_serial',
            'turma': 'id_class',
            'classe': 'id_class',
            'data de matrícula': 'enrollment_date',
            'data de matricula': 'enrollment_date',
            'data': 'enrollment_date',
            'status': 'status',
            'situação': 'status',
            'situacao': 'status',
        }

        # Mapeamento de status (português -> inglês)
        STATUS_MAPPING = {
            'matriculado': 'enrolled',
            'transferido': 'transferred',
            'formado': 'graduated',
            'desistente': 'dropped',
            'ativo': 'active',
            'inativo': 'inactive',
        }

        def normalize_column_name(column):
            """Normaliza o nome da coluna para o padrão em inglês"""
            if not column:
                return None

            # Remove espaços extras e converte para minúsculas
            normalized = column.strip().lower()

            # Se já está em inglês, retorna
            if normalized in ['student_name', 'student_serial', 'id_class', 'enrollment_date', 'status']:
                return normalized

            # Busca no mapeamento
            return COLUMN_MAPPING.get(normalized, normalized)

        def normalize_status(status_value):
            """Normaliza o status para o padrão em inglês"""
            if not status_value:
                return 'enrolled'  # Padrão

            # Remove espaços extras e converte para minúsculas
            normalized = str(status_value).strip().lower()

            # Se já está em inglês, retorna
            if normalized in ['enrolled', 'transferred', 'graduated', 'dropped', 'active', 'inactive']:
                return normalized

            # Busca no mapeamento PT -> EN
            return STATUS_MAPPING.get(normalized, normalized)

        def find_class_by_name_or_id(value, school_id=None):
            """Busca turma por nome ou ID, opcionalmente filtrando por escola"""
            if not value:
                return None, "Valor da turma não fornecido"

            # Remove espaços extras
            value = str(value).strip()

            # Base queryset - filtra por escola se fornecido
            base_queryset = TbClass.objects.all()
            if school_id:
                base_queryset = base_queryset.filter(id_school_id=school_id)

            # Tenta primeiro como ID numérico
            try:
                class_id = int(value)
                try:
                    return base_queryset.get(id=class_id), None
                except TbClass.DoesNotExist:
                    pass  # Continua para tentar por nome
            except (ValueError, TypeError):
                pass  # Não é um número, tenta por nome

            # Tenta buscar por nome exato
            try:
                return base_queryset.get(class_name__iexact=value), None
            except TbClass.DoesNotExist:
                pass

            # Tenta buscar por nome parcial (case insensitive)
            classes = base_queryset.filter(class_name__icontains=value)
            if classes.count() == 1:
                return classes.first(), None
            elif classes.count() > 1:
                class_names = ", ".join([c.class_name for c in classes[:3]])
                return None, f"Múltiplas turmas encontradas para '{value}': {class_names}"

            # Tenta extrair ano e turma (ex: "5º Ano A", "ano: 5º ano")
            match = re.search(r'(\d+)[º°]?\s*(?:ano)?', value.lower())
            if match:
                grade_num = match.group(1)
                # Busca por grade que contenha esse número
                classes = base_queryset.filter(
                    Q(grade__icontains=f"{grade_num}º") |
                    Q(grade__icontains=f"{grade_num}°") |
                    Q(class_name__icontains=value)
                )
                if classes.count() == 1:
                    return classes.first(), None
                elif classes.count() > 1:
                    class_names = ", ".join([c.class_name for c in classes[:3]])
                    return None, f"Múltiplas turmas encontradas: {class_names}"

            school_msg = f" na escola selecionada" if school_id else ""
            return None, f"Turma '{value}' não encontrada{school_msg}"

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Nenhum arquivo enviado'}, status=status.HTTP_400_BAD_REQUEST)

        # Obtém o school_id do request (opcional)
        school_id = request.data.get('school_id')
        if school_id:
            try:
                school_id = int(school_id)
                # Valida se a escola existe
                try:
                    TbSchool.objects.get(id=school_id)
                except TbSchool.DoesNotExist:
                    return Response(
                        {'error': f'Escola com ID {school_id} não encontrada'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except (ValueError, TypeError):
                return Response(
                    {'error': 'ID da escola inválido'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Verifica extensão do arquivo
        file_extension = file.name.split('.')[-1].lower()
        if file_extension not in ['csv', 'xlsx', 'xls']:
            return Response(
                {'error': 'Formato de arquivo inválido. Use CSV ou Excel (.xlsx, .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Processa arquivo CSV
            if file_extension == 'csv':
                 # Try multiple encodings to handle different file formats
                file_content = file.read()
                decoded_file = None
                encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
 
                for encoding in encodings_to_try:
                    try:
                        decoded_file = file_content.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue
 
                if decoded_file is None:
                    return Response(
                        {'error': 'Erro ao processar arquivo: codificação não suportada. Por favor, salve o arquivo como UTF-8.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                reader = csv.DictReader(io.StringIO(decoded_file))
                rows = list(reader)
            # Processa arquivo Excel
            else:
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active
                    headers = [cell.value for cell in sheet[1]]
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    return Response(
                        {'error': 'Biblioteca openpyxl não instalada. Use arquivo CSV.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if not rows:
                return Response({'error': 'Arquivo vazio'}, status=status.HTTP_400_BAD_REQUEST)

            # Normaliza os nomes das colunas na primeira linha
            normalized_rows = []
            for row in rows:
                normalized_row = {}
                for key, value in row.items():
                    normalized_key = normalize_column_name(key)
                    if normalized_key:
                        normalized_row[normalized_key] = value
                normalized_rows.append(normalized_row)

            # Valida e processa os dados
            students_created = []
            students_updated = []
            errors = []
            missing_classes = set()  # Rastreia turmas não encontradas

            with transaction.atomic():
                for idx, row in enumerate(normalized_rows, start=2):
                    try:
                        # Validação dos campos obrigatórios
                        student_name = row.get('student_name', '').strip() if row.get('student_name') else ''
                        student_serial = row.get('student_serial')
                        class_value = row.get('id_class')

                        if not student_name:
                            errors.append(f"Linha {idx}: Nome do aluno é obrigatório")
                            continue

                        if not student_serial:
                            errors.append(f"Linha {idx}: Matrícula é obrigatória")
                            continue

                        if not class_value:
                            errors.append(f"Linha {idx}: Turma é obrigatória")
                            continue

                        # Converte matrícula para inteiro
                        try:
                            student_serial = int(student_serial)
                        except (ValueError, TypeError):
                            errors.append(f"Linha {idx}: Matrícula deve ser um número")
                            continue

                        # Busca a turma por nome ou ID (filtrando por escola se fornecido)
                        class_obj, error = find_class_by_name_or_id(class_value, school_id)
                        if not class_obj:
                            errors.append(f"Linha {idx} ({student_name}): {error}")
                            missing_classes.add(str(class_value))
                            continue

                        # Processa campos opcionais
                        enrollment_date = row.get('enrollment_date')
                        if enrollment_date:
                            if isinstance(enrollment_date, str):
                                try:
                                    enrollment_date = datetime.strptime(enrollment_date, '%Y-%m-%d').date()
                                except ValueError:
                                    errors.append(f"Linha {idx}: Data de matrícula inválida (use YYYY-MM-DD)")
                                    continue
                        else:
                            enrollment_date = datetime.now().date()

                        # Normaliza o status (aceita PT ou EN)
                        student_status = normalize_status(row.get('status'))

                        # Verifica se o aluno já existe (por matrícula)
                        student, created = TbStudents.objects.update_or_create(
                            student_serial=student_serial,
                            defaults={
                                'student_name': student_name,
                                'id_class': class_obj,
                                'enrollment_date': enrollment_date,
                                'status': student_status
                            }
                        )

                        if created:
                            students_created.append(student.student_name)
                        else:
                            students_updated.append(student.student_name)

                    except Exception as e:
                        errors.append(f"Linha {idx}: {str(e)}")
                        continue

            # Se houver turmas não encontradas, lista as turmas disponíveis (filtradas por escola se fornecido)
            available_classes = None
            if missing_classes:
                classes_query = TbClass.objects.all()
                if school_id:
                    classes_query = classes_query.filter(id_school_id=school_id)
                all_classes = classes_query.values('id', 'class_name', 'grade')
                available_classes = [
                    {
                        'id': c['id'],
                        'name': c['class_name'],
                        'grade': c['grade']
                    }
                    for c in all_classes
                ]

            response_data = {
                'success': True,
                'message': f'{len(students_created)} alunos criados, {len(students_updated)} alunos atualizados',
                'created': len(students_created),
                'updated': len(students_updated),
                'errors': errors if errors else None
            }

            # Adiciona informações extras quando houver erros
            if missing_classes:
                response_data['missing_classes'] = list(missing_classes)
                response_data['available_classes'] = available_classes
                school_filter_msg = f" na escola selecionada" if school_id else ""
                response_data['suggestion'] = f'Algumas turmas não foram encontradas{school_filter_msg}. Verifique se as turmas existem no sistema antes de importar os alunos.'

            return Response(response_data, status=status.HTTP_201_CREATED if students_created else status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class TbSubjectViewSet(viewsets.ModelViewSet):
    """Disciplinas (Subjects)"""
    queryset = TbSubject.objects.all()
    serializer_class = TbSubjectSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['subject_name', 'description']
    ordering_fields = ['subject_name']

    
# ============================================
# VIEWSETS DE INDICADORES IDEB
# ============================================

class TbSchoolIdebIndicatorsViewSet(viewsets.ModelViewSet):
    """Indicadores IDEB das Escolas"""
    queryset = TbSchoolIdebIndicators.objects.all().select_related('id_school')
    serializer_class = TbSchoolIdebIndicatorsSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_school', 'fiscal_year']


class TbClassIdebIndicatorsViewSet(viewsets.ModelViewSet):
    """Indicadores IDEB das Turmas"""
    queryset = TbClassIdebIndicators.objects.all().select_related('id_class')
    serializer_class = TbClassIdebIndicatorsSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_class', 'fiscal_year']


# ============================================
# VIEWSETS DE COMPETÊNCIAS E DESCRITORES
# ============================================

class TbCompetencyIdebViewSet(viewsets.ModelViewSet):
    """Competências IDEB"""
    queryset = TbCompetencyIdeb.objects.all()
    serializer_class = TbCompetencyIdebSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['subject', 'grade']
    search_fields = ['competency_code', 'competency_name']


class TbDescriptorsCatalogViewSet(viewsets.ModelViewSet):
    """Catálogo de Descritores"""
    queryset = TbDescriptorsCatalog.objects.all()
    serializer_class = TbDescriptorsCatalogSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['subject', 'grade', 'learning_field']
    search_fields = ['descriptor_code', 'descriptor_name', 'descriptor_description']


# ============================================
# VIEWSETS DE EXAMES E QUESTÕES
# ============================================

class TbExamsViewSet(viewsets.ModelViewSet):
    """Exames com funcionalidades extras"""
    queryset = TbExams.objects.all()
    serializer_class = TbExamsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['subject', 'school_year']
    search_fields = ['exam_code', 'exam_name']
    ordering_fields = ['created_at', 'exam_name']

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return TbExamsDetailSerializer
        return TbExamsSerializer

    @action(detail=True, methods=['get'])
    def questions(self, request, pk=None):
        """Lista todas as questões do exame com alternativas"""
        exam = self.get_object()
        questions = TbQuestions.objects.filter(id_exam=exam).select_related('id_descriptor')
        serializer = TbQuestionsSerializer(questions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Estatísticas do exame"""
        exam = self.get_object()

        applications = TbExamApplications.objects.filter(id_exam=exam)
        results = TbExamResults.objects.filter(id_exam_application__id_exam=exam)

        stats = {
            'total_applications': applications.count(),
            'total_students': results.count(),
            'average_score': results.aggregate(Avg('total_score'))['total_score__avg'] or 0,
            'highest_score': results.aggregate(Max('total_score'))['total_score__max'] or 0,
            'lowest_score': results.aggregate(Min('total_score'))['total_score__min'] or 0,
            'total_questions': TbQuestions.objects.filter(id_exam=exam).count(),
        }

        return Response(stats)

    @action(detail=True, methods=['post'])
    def upload_file(self, request, pk=None):
        """
        Faz upload do arquivo da prova (PDF, imagem, etc) para S3 ou storage local
        """
        from django.conf import settings
        import boto3
        from botocore.exceptions import ClientError
        import uuid

        exam = self.get_object()
        file = request.FILES.get('file')

        if not file:
            return Response(
                {'error': 'Nenhum arquivo enviado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tipo de arquivo
        allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx']
        file_extension = '.' + file.name.split('.')[-1].lower()

        if file_extension not in allowed_extensions:
            return Response(
                {
                    'error': f'Tipo de arquivo não permitido. Use: {", ".join(allowed_extensions)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tamanho (max 50MB)
        max_size = 50 * 1024 * 1024  # 50MB
        if file.size > max_size:
            return Response(
                {'error': 'Arquivo muito grande. Tamanho máximo: 50MB'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            if settings.USE_S3:
                # Upload para S3
                s3_client = boto3.client(
                    's3',
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME
                )

                # Gera nome único para o arquivo
                unique_filename = f"{settings.EXAM_FILES_LOCATION}/{exam.exam_code}_{uuid.uuid4()}{file_extension}"

                # Faz upload
                s3_client.upload_fileobj(
                    file,
                    settings.AWS_STORAGE_BUCKET_NAME,
                    unique_filename,
                    ExtraArgs={
                        'ContentType': file.content_type,
                        'ACL': 'public-read'
                    }
                )

                # URL do arquivo no S3
                file_url = f"https://{settings.AWS_S3_CUSTOM_DOMAIN}/{unique_filename}"

            else:
                # Upload local (desenvolvimento)
                import os
                from django.core.files.storage import default_storage

                # Cria diretório se não existir
                upload_path = os.path.join(settings.EXAM_FILES_LOCATION, exam.exam_code)
                os.makedirs(os.path.join(settings.MEDIA_ROOT, upload_path), exist_ok=True)

                # Salva arquivo
                unique_filename = f"{upload_path}/{uuid.uuid4()}{file_extension}"
                file_path = default_storage.save(unique_filename, file)

                # URL do arquivo local
                file_url = f"{settings.MEDIA_URL}{file_path}"

            # Atualiza prova com URL do arquivo
            exam.exam_file = file_url
            exam.save()

            return Response({
                'success': True,
                'message': 'Arquivo enviado com sucesso',
                'file_url': file_url,
                'file_name': file.name,
                'file_size': file.size
            }, status=status.HTTP_200_OK)

        except ClientError as e:
            return Response(
                {'error': f'Erro ao fazer upload para S3: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def import_answer_key(self, request):
        """
        Importa gabarito completo da prova via CSV/Excel

        Formato esperado:
        - codigo_prova: Código único da prova (ex: PROVA2024_MAT_5)
        - nome_prova: Nome da prova
        - disciplina: Disciplina (Matemática, Português, etc)
        - ano_escolar: Ano escolar (1-9)
        - numero_questao: Número da questão (1, 2, 3...)
        - resposta_correta: Letra da resposta (A, B, C, D, E)
        - codigo_descritor: Código do descritor (ex: D01, D02)
        - pontos: Pontos da questão (ex: 1.0, 1.5)
        - dificuldade: Dificuldade (easy, medium, hard)
        - enunciado: Texto da questão (opcional)
        """
        import csv
        import io
        from decimal import Decimal

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Nenhum arquivo enviado'}, status=status.HTTP_400_BAD_REQUEST)

        # Verifica extensão
        file_extension = file.name.split('.')[-1].lower()
        if file_extension not in ['csv', 'xlsx', 'xls']:
            return Response(
                {'error': 'Formato inválido. Use CSV ou Excel (.xlsx, .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Processa arquivo
            if file_extension == 'csv':
                 # Try multiple encodings to handle different file formats
                file_content = file.read()
                decoded_file = None
                encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
 
                for encoding in encodings_to_try:
                    try:
                        decoded_file = file_content.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue
 
                if decoded_file is None:
                    return Response(
                        {'error': 'Erro ao processar arquivo: codificação não suportada. Por favor, salve o arquivo como UTF-8.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                reader = csv.DictReader(io.StringIO(decoded_file))
                rows = list(reader)
            else:
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active
                    headers = [cell.value for cell in sheet[1]]
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    return Response(
                        {'error': 'Biblioteca openpyxl não instalada. Use CSV.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if not rows:
                return Response({'error': 'Arquivo vazio'}, status=status.HTTP_400_BAD_REQUEST)

            # Agrupa questões por código de prova
            exams_data = {}
            errors = []
            descriptors_not_found = set()

            for idx, row in enumerate(rows, start=2):
                try:
                    codigo_prova = str(row.get('codigo_prova', '')).strip()
                    nome_prova = str(row.get('nome_prova', '')).strip()
                    disciplina = str(row.get('disciplina', '')).strip()
                    ano_escolar = row.get('ano_escolar')
                    numero_questao = row.get('numero_questao')
                    resposta_correta = str(row.get('resposta_correta', '')).strip().upper()
                    codigo_descritor = str(row.get('codigo_descritor', '')).strip()
                    pontos = row.get('pontos', 1.0)
                    dificuldade = str(row.get('dificuldade', 'medium')).strip().lower()
                    enunciado = str(row.get('enunciado', '')).strip()

                    # Validações
                    if not codigo_prova:
                        errors.append(f"Linha {idx}: codigo_prova obrigatório")
                        continue
                    if not nome_prova:
                        errors.append(f"Linha {idx}: nome_prova obrigatório")
                        continue
                    if not numero_questao:
                        errors.append(f"Linha {idx}: numero_questao obrigatório")
                        continue
                    if not resposta_correta or resposta_correta not in ['A', 'B', 'C', 'D', 'E']:
                        errors.append(f"Linha {idx}: resposta_correta deve ser A, B, C, D ou E")
                        continue

                    # Converte tipos
                    try:
                        numero_questao = int(numero_questao)
                        pontos = Decimal(str(pontos))
                        if ano_escolar:
                            ano_escolar = str(ano_escolar).strip()
                    except (ValueError, TypeError) as e:
                        errors.append(f"Linha {idx}: Erro de conversão - {str(e)}")
                        continue

                    # Valida descritor
                    descriptor = None
                    if codigo_descritor:
                        try:
                            descriptor = TbDescriptorsCatalog.objects.get(descriptor_code=codigo_descritor)
                        except TbDescriptorsCatalog.DoesNotExist:
                            descriptors_not_found.add(codigo_descritor)

                    # Agrupa por prova
                    if codigo_prova not in exams_data:
                        exams_data[codigo_prova] = {
                            'nome_prova': nome_prova,
                            'disciplina': disciplina,
                            'ano_escolar': ano_escolar,
                            'questions': []
                        }

                    exams_data[codigo_prova]['questions'].append({
                        'numero_questao': numero_questao,
                        'resposta_correta': resposta_correta,
                        'descriptor': descriptor,
                        'pontos': pontos,
                        'dificuldade': dificuldade,
                        'enunciado': enunciado
                    })

                except Exception as e:
                    errors.append(f"Linha {idx}: {str(e)}")
                    continue

            # Cria provas e questões
            created_exams = []
            with transaction.atomic():
                for codigo_prova, exam_data in exams_data.items():
                    # Verifica se prova já existe
                    exam, created = TbExams.objects.get_or_create(
                        exam_code=codigo_prova,
                        defaults={
                            'exam_name': exam_data['nome_prova'],
                            'subject': exam_data['disciplina'],
                            'school_year': exam_data['ano_escolar'],
                            'total_questions': len(exam_data['questions'])
                        }
                    )

                    if not created:
                        # Atualiza informações da prova se já existe
                        exam.exam_name = exam_data['nome_prova']
                        exam.subject = exam_data['disciplina']
                        exam.school_year = exam_data['ano_escolar']
                        exam.total_questions = len(exam_data['questions'])
                        exam.save()

                        # Remove alternativas e questões antigas

                        questions_to_delete = TbQuestions.objects.filter(id_exam=exam)

                        for question in questions_to_delete:

                            TbAlternatives.objects.filter(id_question=question).delete()

                        questions_to_delete.delete()

                    # Cria questões
                    for q_data in exam_data['questions']:
                        question = TbQuestions.objects.create(
                            id_exam=exam,
                            question_number=q_data['numero_questao'],
                            question_text=q_data['enunciado'] or f"Questão {q_data['numero_questao']}",
                            question_type='multiple_choice',
                            correct_answer=q_data['resposta_correta'],
                            difficulty_level=q_data['dificuldade'],
                            points=q_data['pontos'],
                            id_descriptor=q_data['descriptor']
                        )

                        # Cria alternativas genéricas (A, B, C, D, E)
                        for i, letter in enumerate(['A', 'B', 'C', 'D', 'E'], start=1):
                            TbAlternatives.objects.create(
                                id_question=question,
                                alternative_order=i,
                                alternative_text=f"Alternativa {letter}",
                                is_correct=(letter == q_data['resposta_correta'])
                            )

                    created_exams.append({
                        'codigo': codigo_prova,
                        'nome': exam_data['nome_prova'],
                        'questoes': len(exam_data['questions']),
                        'criada': created
                    })

            response_data = {
                'success': True,
                'message': f'{len(created_exams)} prova(s) importada(s)',
                'exams': created_exams,
                'errors': errors if errors else None
            }

            if descriptors_not_found:
                response_data['warning'] = 'Alguns descritores não foram encontrados'
                response_data['descriptors_not_found'] = list(descriptors_not_found)

            return Response(response_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class TbQuestionsViewSet(viewsets.ModelViewSet):
    """Questões"""
    queryset = TbQuestions.objects.all().select_related('id_exam', 'id_descriptor')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['id_exam', 'difficulty_level', 'id_descriptor']
    search_fields = ['question_text']
    
    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return TbQuestionsCreateSerializer
        return TbQuestionsSerializer


class TbAlternativesViewSet(viewsets.ModelViewSet):
    """Alternativas"""
    queryset = TbAlternatives.objects.all().select_related('id_question')
    serializer_class = TbAlternativesSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_question', 'is_correct']


class TbQuestionCompetencyViewSet(viewsets.ModelViewSet):
    """Relação Questão-Competência"""
    queryset = TbQuestionCompetency.objects.all().select_related('id_question', 'id_competency')
    serializer_class = TbQuestionCompetencySerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_question', 'id_competency']


# ============================================
# VIEWSETS DE APLICAÇÕES E RESULTADOS
# ============================================

class TbExamApplicationsViewSet(viewsets.ModelViewSet):
    """Aplicações de Exames"""
    queryset = TbExamApplications.objects.all().select_related(
        'id_exam',
        'id_class',
        'id_class__id_school',
        'id_teacher'
    )
    serializer_class = TbExamApplicationsSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = [
        'id_exam', 'id_class', 'id_teacher', 'status',
        'application_type', 'fiscal_year'
    ]
    ordering_fields = ['application_date', 'created_at']

    def get_serializer_class(self):
        # Detail endpoint usa serializer detalhado
        if self.action == 'retrieve' or self.action == 'list':
            return TbExamApplicationsDetailSerializer
        return TbExamApplicationsSerializer

    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """Lista alunos da turma que devem fazer a prova"""
        application = self.get_object()
        students = TbStudents.objects.filter(id_class=application.id_class)
        
        # Adiciona info se já tem resultado
        students_data = []
        for student in students:
            has_result = TbExamResults.objects.filter(
                id_student=student,
                id_exam_application=application
            ).exists()
            
            students_data.append({
                'id_student': student.id_student,
                'student_serial': student.student_serial,
                'student_name': student.student_name,
                'status': student.status,
                'has_result': has_result
            })
        
        return Response(students_data)

    @action(detail=True, methods=['get'])
    def results(self, request, pk=None):
        """Retorna todos os resultados de uma aplicação"""
        application = self.get_object()
        results = TbExamResults.objects.filter(
            id_exam_application=application
        ).select_related('id_student')
        serializer = TbExamResultsDetailSerializer(results, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        """Altera o status da aplicação"""
        application = self.get_object()
        new_status = request.data.get('status')

        if new_status not in ['scheduled', 'in_progress', 'completed', 'cancelled']:
            return Response(
                {'error': 'Status inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        application.status = new_status
        application.save()

        serializer = self.get_serializer(application)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def import_student_answers(self, request):
        """
        Importa respostas dos alunos via CSV/Excel

        Formato esperado:
        - codigo_prova: Código da prova previamente importada
        - id_turma: ID da turma
        - matricula_aluno: Matrícula do aluno
        - q1, q2, q3...: Respostas (A, B, C, D, E ou vazio para questão em branco)

        Exemplo CSV:
        codigo_prova,id_turma,matricula_aluno,q1,q2,q3,q4,q5
        PROVA2024_MAT_5,1,12345,A,C,B,D,A
        PROVA2024_MAT_5,1,67890,B,C,A,D,
        """
        import csv
        import io
        from datetime import datetime

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Nenhum arquivo enviado'}, status=status.HTTP_400_BAD_REQUEST)

        # Verifica extensão
        file_extension = file.name.split('.')[-1].lower()
        if file_extension not in ['csv', 'xlsx', 'xls']:
            return Response(
                {'error': 'Formato inválido. Use CSV ou Excel (.xlsx, .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Processa arquivo
            if file_extension == 'csv':
                decoded_file = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded_file))
                rows = list(reader)
            else:
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active
                    headers = [cell.value for cell in sheet[1]]
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    return Response(
                        {'error': 'Biblioteca openpyxl não instalada. Use CSV.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if not rows:
                return Response({'error': 'Arquivo vazio'}, status=status.HTTP_400_BAD_REQUEST)

            # Agrupa por código de prova e turma
            applications_data = {}
            errors = []
            students_not_found = set()
            exams_not_found = set()

            for idx, row in enumerate(rows, start=2):
                try:
                    codigo_prova = str(row.get('codigo_prova', '')).strip()
                    id_turma = row.get('id_turma')
                    matricula_aluno = row.get('matricula_aluno')

                    # Validações básicas
                    if not codigo_prova:
                        errors.append(f"Linha {idx}: codigo_prova obrigatório")
                        continue
                    if not id_turma:
                        errors.append(f"Linha {idx}: id_turma obrigatório")
                        continue
                    if not matricula_aluno:
                        errors.append(f"Linha {idx}: matricula_aluno obrigatório")
                        continue

                    # Converte tipos
                    try:
                        id_turma = int(id_turma)
                        matricula_aluno = int(matricula_aluno)
                    except (ValueError, TypeError):
                        errors.append(f"Linha {idx}: id_turma e matricula_aluno devem ser números")
                        continue

                    # Busca prova
                    try:
                        exam = TbExams.objects.get(exam_code=codigo_prova)
                    except TbExams.DoesNotExist:
                        exams_not_found.add(codigo_prova)
                        errors.append(f"Linha {idx}: Prova '{codigo_prova}' não encontrada")
                        continue

                    # Busca turma
                    try:
                        class_obj = TbClass.objects.get(id=id_turma)
                    except TbClass.DoesNotExist:
                        errors.append(f"Linha {idx}: Turma ID {id_turma} não encontrada")
                        continue

                    # Busca aluno
                    try:
                        student = TbStudents.objects.get(
                            student_serial=matricula_aluno,
                            id_class=class_obj
                        )
                    except TbStudents.DoesNotExist:
                        students_not_found.add(matricula_aluno)
                        errors.append(f"Linha {idx}: Aluno matrícula {matricula_aluno} não encontrado na turma")
                        continue

                    # Extrai respostas (q1, q2, q3...)
                    answers = {}
                    for key, value in row.items():
                        if key.startswith('q') and key[1:].isdigit():
                            question_number = int(key[1:])
                            answer_value = str(value).strip().upper() if value else ''
                            if answer_value and answer_value not in ['A', 'B', 'C', 'D', 'E']:
                                errors.append(f"Linha {idx}, {key}: Resposta deve ser A, B, C, D, E ou vazio")
                                continue
                            answers[question_number] = answer_value

                    if not answers:
                        errors.append(f"Linha {idx}: Nenhuma resposta encontrada (colunas q1, q2, q3...)")
                        continue

                    # Agrupa por aplicação (prova + turma)
                    app_key = (codigo_prova, id_turma)
                    if app_key not in applications_data:
                        applications_data[app_key] = {
                            'exam': exam,
                            'class': class_obj,
                            'students_answers': []
                        }

                    applications_data[app_key]['students_answers'].append({
                        'student': student,
                        'answers': answers
                    })

                except Exception as e:
                    errors.append(f"Linha {idx}: {str(e)}")
                    continue

            # Processa aplicações e respostas
            processed_students = 0
            created_applications = []

            with transaction.atomic():
                for (codigo_prova, id_turma), app_data in applications_data.items():
                    exam = app_data['exam']
                    class_obj = app_data['class']

                    # Cria ou busca aplicação
                    application, created = TbExamApplications.objects.get_or_create(
                        id_exam=exam,
                        id_class=class_obj,
                        defaults={
                            'application_date': datetime.now().date(),
                            'status': 'completed',
                            'fiscal_year': datetime.now().year
                        }
                    )

                    if created:
                        created_applications.append({
                            'exam': exam.exam_name,
                            'class': class_obj.class_name
                        })

                    # Busca questões da prova
                    questions = TbQuestions.objects.filter(id_exam=exam).select_related('id_descriptor')
                    questions_dict = {q.question_number: q for q in questions}

                    # Processa respostas de cada aluno
                    for student_data in app_data['students_answers']:
                        student = student_data['student']
                        answers = student_data['answers']

                        # Verifica se aluno já fez a prova
                        if TbStudentAnswers.objects.filter(
                            id_student=student,
                            id_exam_application=application
                        ).exists():
                            errors.append(
                                f"Aluno {student.student_name} (mat: {student.student_serial}) "
                                f"já tem respostas para esta prova"
                            )
                            continue

                        # Cria respostas
                        student_answers = []
                        for q_number, answer_letter in answers.items():
                            if q_number not in questions_dict:
                                continue

                            question = questions_dict[q_number]

                            # Busca alternativa selecionada
                            selected_alternative = None
                            is_correct = False

                            if answer_letter:  # Se não está em branco
                                # Converte letra para número (A=1, B=2, etc)
                                alternative_order = ord(answer_letter) - ord('A') + 1
                                try:
                                    selected_alternative = TbAlternatives.objects.get(
                                        id_question=question,
                                        alternative_order=alternative_order
                                    )
                                    is_correct = selected_alternative.is_correct
                                except TbAlternatives.DoesNotExist:
                                    pass

                            # Cria resposta do aluno
                            student_answer = TbStudentAnswers.objects.create(
                                id_student=student,
                                id_exam_application=application,
                                id_question=question,
                                id_selected_alternative=selected_alternative,
                                answer_text=answer_letter,
                                is_correct=is_correct
                            )
                            student_answers.append(student_answer)

                        # Calcula resultado
                        if student_answers:
                            self._calculate_result_and_descriptors(student, application, student_answers)
                            processed_students += 1

            response_data = {
                'success': True,
                'message': f'Respostas de {processed_students} aluno(s) importadas com sucesso',
                'processed_students': processed_students,
                'created_applications': created_applications,
                'errors': errors if errors else None
            }

            if students_not_found:
                response_data['students_not_found'] = list(students_not_found)

            if exams_not_found:
                response_data['exams_not_found'] = list(exams_not_found)

            return Response(response_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _calculate_result_and_descriptors(self, student, application, answers):
        """Calcula resultado e atribui descritores conquistados"""
        # Calcula pontuação
        total_score = sum([ans.id_question.points for ans in answers if ans.is_correct])
        max_score = sum([ans.id_question.points for ans in answers])
        correct_count = sum([1 for ans in answers if ans.is_correct])
        wrong_count = sum([1 for ans in answers if not ans.is_correct and ans.answer_text])
        blank_count = sum([1 for ans in answers if not ans.answer_text])

        # Cria resultado
        TbExamResults.objects.update_or_create(
            id_student=student,
            id_exam_application=application,
            defaults={
                'total_score': total_score,
                'max_score': max_score,
                'correct_answers': correct_count,
                'wrong_answers': wrong_count,
                'blank_answers': blank_count
            }
        )

        # Atribui descritores das questões corretas
        for answer in answers:
            if answer.is_correct and answer.id_question.id_descriptor:
                TbStudentDescriptorAchievements.objects.get_or_create(
                    id_student=student,
                    id_descriptor=answer.id_question.id_descriptor,
                    defaults={'id_exam_application': application}
                )


class TbAssessmentMetadataViewSet(viewsets.ModelViewSet):
    """Metadados de Avaliações"""
    queryset = TbAssessmentMetadata.objects.all().select_related('id_exam_application')
    serializer_class = TbAssessmentMetadataSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['application_type', 'assessment_period', 'fiscal_year']


class TbStudentAnswersViewSet(viewsets.ModelViewSet):
    """Respostas dos Alunos"""
    queryset = TbStudentAnswers.objects.all().select_related(
        'id_student', 'id_question', 'id_selected_alternative'
    )
    serializer_class = TbStudentAnswersSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_student', 'id_exam_application', 'id_question', 'is_correct']

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Lançamento em lote de respostas"""
        serializer = BulkStudentAnswersSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated_data = serializer.validated_data
        id_student = validated_data['id_student']
        id_exam_application = validated_data['id_exam_application']
        answers = validated_data['answers']

        # Verificar se o aluno já fez esta prova
        existing_answers = TbStudentAnswers.objects.filter(
            id_student_id=id_student,
            id_exam_application_id=id_exam_application
        ).exists()

        if existing_answers:
            return Response({
                'error': 'Este aluno já realizou esta prova'
            }, status=status.HTTP_400_BAD_REQUEST)

        created_answers = []
        errors = []

        try:
            with transaction.atomic():
                for answer in answers:
                    try:
                        question = TbQuestions.objects.get(id=answer['id_question'])
                        selected_alt_id = answer.get('id_selected_alternative')

                        # Verificar se a resposta está correta
                        is_correct = False
                        if selected_alt_id:
                            # Buscar pela alternativa marcada como correta
                            correct_alt = TbAlternatives.objects.filter(
                                id_question=question,
                                is_correct=True
                            ).first()

                            if correct_alt:
                                is_correct = int(selected_alt_id) == correct_alt.id

                        student_answer = TbStudentAnswers.objects.create(
                            id_student_id=id_student,
                            id_exam_application_id=id_exam_application,
                            id_question_id=answer['id_question'],
                            id_selected_alternative_id=selected_alt_id,
                            answer_text=answer.get('answer_text', ''),
                            is_correct=is_correct
                        )
                        created_answers.append(student_answer)
                    except TbQuestions.DoesNotExist:
                        errors.append(f"Questão {answer.get('id_question')} não encontrada")
                        continue
                    except Exception as e:
                        errors.append(f"Erro na questão {answer.get('id_question')}: {str(e)}")
                        continue

                # Calcular e criar resultado automático
                if created_answers:
                    self._calculate_exam_result(id_student, id_exam_application)

        except Exception as e:
            return Response({
                'error': f'Erro ao processar respostas: {str(e)}',
                'details': errors
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'message': 'Respostas registradas com sucesso',
            'total_answers': len(created_answers),
            'errors': errors if errors else None
        }, status=status.HTTP_201_CREATED)
    
    def _calculate_exam_result(self, id_student, id_exam_application):
        """Calcula resultado automático baseado nas respostas"""
        answers = TbStudentAnswers.objects.filter(
            id_student_id=id_student,
            id_exam_application_id=id_exam_application
        ).select_related('id_question')
        
        total_score = sum([
            ans.id_question.points for ans in answers if ans.is_correct
        ])
        
        max_score = sum([ans.id_question.points for ans in answers])
        
        correct_answers = answers.filter(is_correct=True).count()
        wrong_answers = answers.filter(is_correct=False).count()
        blank_answers = 0
        
        # Criar ou atualizar resultado
        TbExamResults.objects.update_or_create(
            id_student_id=id_student,
            id_exam_application_id=id_exam_application,
            defaults={
                'total_score': total_score,
                'max_score': max_score,
                'correct_answers': correct_answers,
                'wrong_answers': wrong_answers,
                'blank_answers': blank_answers
            }
        )


class TbExamResultsViewSet(viewsets.ModelViewSet):
    """Resultados dos Exames"""
    queryset = TbExamResults.objects.all().select_related(
        'id_student',
        'id_student__id_class',
        'id_exam_application',
        'id_exam_application__id_exam'
    )
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['id_student', 'id_exam_application']
    ordering_fields = ['total_score', 'created_at']

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return TbExamResultsDetailSerializer
        return TbExamResultsSerializer

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Criação em lote de resultados"""
        results_data = request.data if isinstance(request.data, list) else [request.data]
        
        created_results = []
        errors = []
        
        with transaction.atomic():
            for data in results_data:
                try:
                    serializer = TbExamResultsSerializer(data=data)
                    if serializer.is_valid():
                        result = serializer.save()
                        created_results.append(result)
                    else:
                        errors.append({
                            'data': data,
                            'errors': serializer.errors
                        })
                except Exception as e:
                    errors.append({
                        'data': data,
                        'errors': str(e)
                    })
        
        return Response({
            'created': len(created_results),
            'errors': errors
        }, status=status.HTTP_201_CREATED if created_results else status.HTTP_400_BAD_REQUEST)


# ============================================
# VIEWSETS DE PROGRESSO E CONQUISTAS
# ============================================

class TbStudentDescriptorAchievementsViewSet(viewsets.ModelViewSet):
    """Conquistas de Descritores dos Alunos"""
    queryset = TbStudentDescriptorAchievements.objects.all().select_related(
        'id_student', 'id_descriptor'
    )
    serializer_class = TbStudentDescriptorAchievementsSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['id_student', 'id_descriptor', 'id_exam_application']
    ordering_fields = ['achieved_at']


class TbStudentLearningProgressViewSet(viewsets.ModelViewSet):
    """Progresso de Aprendizagem dos Alunos"""
    queryset = TbStudentLearningProgress.objects.all().select_related(
        'id_student', 'id_descriptor', 'id_exam_application'
    )
    serializer_class = TbStudentLearningProgressSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['id_student', 'id_descriptor', 'id_exam_application']
    ordering_fields = ['assessment_date', 'descriptor_mastery']

    @action(detail=False, methods=['get'])
    def by_student(self, request):
        """Retorna o progresso de um aluno específico"""
        student_id = request.query_params.get('student_id')
        if not student_id:
            return Response({'error': 'student_id is required'}, status=400)

        progress = self.queryset.filter(id_student=student_id).order_by('-assessment_date')
        serializer = self.get_serializer(progress, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def low_performance(self, request):
        """Retorna alunos com baixo desempenho (< 50%)"""
        threshold = request.query_params.get('threshold', 50)
        progress = self.queryset.filter(descriptor_mastery__lt=threshold)
        serializer = self.get_serializer(progress, many=True)
        return Response(serializer.data)
    
class StudentProfileViewSet(viewsets.ReadOnlyModelViewSet):
    """Perfil completo do aluno com descritores conquistados"""
    queryset = TbStudents.objects.all().select_related(
        'id_class',
        'id_class__id_school',
        'id_class__id_teacher'
    )
    serializer_class = TbStudentsSerializer
    lookup_field = 'id_student'  # ✅ Importante: usar id_student como lookup

    def retrieve(self, request, id_student=None):
        """Retorna perfil completo do aluno (sobrescreve o retrieve padrão)"""
        return self._get_profile_data()

    def _get_profile_data(self):
        """Método auxiliar que retorna os dados completos do perfil do aluno"""
        try:
            # ✅ Usar self.get_object() para buscar corretamente
            student = self.get_object()
            # Dados básicos do aluno com tratamento de None
            profile_data = {
                'id_student': student.id_student,
                'student_serial': student.student_serial,
                'student_name': student.student_name,
                'status': student.status,
                'enrollment_date': student.enrollment_date,
                'class_name': student.id_class.class_name if student.id_class else None,
                'school_name': (student.id_class.id_school.school
                               if student.id_class and student.id_class.id_school else None),
                'grade': student.id_class.grade if student.id_class else None,
                'shift': student.id_class.shift if student.id_class else None,
            }
            # Descritores conquistados
            achievements = TbStudentDescriptorAchievements.objects.filter(
                id_student=student
            ).select_related('id_descriptor', 'id_exam_application')
            descriptors_achieved = []
            for achievement in achievements:
                exam_name = None
                if achievement.id_exam_application and achievement.id_exam_application.id_exam:
                    exam_name = achievement.id_exam_application.id_exam.exam_name
 
                descriptors_achieved.append({
                    'id': achievement.id_descriptor.id,
                    'descriptor_code': achievement.id_descriptor.descriptor_code,
                    'descriptor_name': achievement.id_descriptor.descriptor_name,
                    'subject': achievement.id_descriptor.subject,
                    'learning_field': achievement.id_descriptor.learning_field,
                    'icon': achievement.id_descriptor.icon,
                    'achieved_at': achievement.achieved_at,
                    'exam_name': exam_name,
                })
            # Todos os descritores do catálogo
            all_descriptors = TbDescriptorsCatalog.objects.all()
            total_descriptors = all_descriptors.count()
            achieved_count = len(descriptors_achieved)
            # Agrupa descritores por disciplina
            descriptors_by_subject = {}
            for desc in all_descriptors:
                subject = desc.subject or 'Outros'
                if subject not in descriptors_by_subject:
                    descriptors_by_subject[subject] = {
                        'total': 0,
                        'achieved': 0,
                        'descriptors': []
                    }
                descriptors_by_subject[subject]['total'] += 1
                descriptors_by_subject[subject]['descriptors'].append({
                    'id': desc.id,
                    'code': desc.descriptor_code,
                    'name': desc.descriptor_name,
                    'achieved': any(d['id'] == desc.id for d in descriptors_achieved)
                })

                # Conta conquistados
                if any(d['id'] == desc.id for d in descriptors_achieved):
                    descriptors_by_subject[subject]['achieved'] += 1

            # Estatísticas de progresso
            learning_progress = TbStudentLearningProgress.objects.filter(
                id_student=student
            ).select_related('id_descriptor').order_by('-assessment_date')[:5]
            recent_progress = []
            for progress in learning_progress:
                recent_progress.append({
                    'descriptor_name': progress.id_descriptor.descriptor_name,
                    'descriptor_code': progress.id_descriptor.descriptor_code,
                    'descriptor_mastery': float(progress.descriptor_mastery),
                    'assessment_date': progress.assessment_date,
                    'score': float(progress.score),
                    'max_score': float(progress.max_score),
                })
            # Últimos resultados de provas
            exam_results = TbExamResults.objects.filter(
                id_student=student
            ).select_related('id_exam_application__id_exam').order_by('-created_at')[:5]
            recent_exams = []
            for result in exam_results:
                recent_exams.append({
                    'exam_name': result.id_exam_application.id_exam.exam_name,
                    'total_score': float(result.total_score) if result.total_score else 0,
                    'max_score': float(result.max_score) if result.max_score else 0,
                    'percentage': round((float(result.total_score or 0) / float(result.max_score or 1)) * 100, 2),
                    'correct_answers': result.correct_answers,
                    'wrong_answers': result.wrong_answers,
                    'application_date': result.id_exam_application.application_date,
                })
            profile_data['descriptors'] = {
                'achieved': descriptors_achieved,
                'total_count': total_descriptors,
                'achieved_count': achieved_count,
                'percentage': round((achieved_count / total_descriptors * 100), 2) if total_descriptors > 0 else 0,
                'by_subject': descriptors_by_subject,
            }

            profile_data['recent_progress'] = recent_progress
            profile_data['recent_exams'] = recent_exams

            return Response(profile_data)

        except TbStudents.DoesNotExist:
            return Response({'error': 'Aluno não encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def profile(self, request, id_student=None):
        """Retorna perfil completo do aluno (mantido por compatibilidade)"""
        return self._get_profile_data()

    @action(detail=True, methods=['post'])
    def toggle_descriptor(self, request, id_student=None):
        """Toggle descritor para o aluno (atribuir/desatribuir)"""
        try:
            student = self.get_object()
            descriptor_id = request.data.get('descriptor_id')

            if not descriptor_id:
                return Response(
                    {'error': 'descriptor_id é obrigatório'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verificar se o descritor existe
            try:
                descriptor = TbDescriptorsCatalog.objects.get(id=descriptor_id)
            except TbDescriptorsCatalog.DoesNotExist:
                return Response(
                    {'error': 'Descritor não encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Verificar se já existe uma conquista
            achievement = TbStudentDescriptorAchievements.objects.filter(
                id_student=student,
                id_descriptor=descriptor
            ).first()

            if achievement:
                # Se existe, remover (desatribuir)
                achievement.delete()
                return Response({
                    'message': 'Descritor desatribuído com sucesso',
                    'achieved': False,
                    'descriptor_id': descriptor_id
                }, status=status.HTTP_200_OK)
            else:
                # Se não existe, criar (atribuir)
                TbStudentDescriptorAchievements.objects.create(
                    id_student=student,
                    id_descriptor=descriptor,
                    id_exam_application=None  # Manual assignment
                )
                return Response({
                    'message': 'Descritor atribuído com sucesso',
                    'achieved': True,
                    'descriptor_id': descriptor_id
                }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar solicitação: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
