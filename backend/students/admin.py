from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.utils.html import format_html
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
from .models import (
    TbStudents, TbClass, TbSchool, TbCity, TbTeacher,
    TbSubject, TbTeacherSubject, TbTeacherSchool
)


@admin.register(TbStudents)
class TbStudentsAdmin(admin.ModelAdmin):
    list_display = ['id_student', 'student_serial', 'student_name', 'id_class', 'status', 'enrollment_date']
    list_filter = ['status', 'id_class', 'enrollment_date']
    search_fields = ['student_name', 'student_serial']
    list_per_page = 50

    # Campos no formulário de edição
    fields = ['student_serial', 'student_name', 'id_class', 'enrollment_date', 'status']

    def get_urls(self):
        """Adiciona URLs customizadas ao admin"""
        urls = super().get_urls()
        custom_urls = [
            path('importar-estudantes/', self.admin_site.admin_view(self.import_students_view), name='students_import'),
            path('download-modelo/', self.admin_site.admin_view(self.download_template), name='students_template'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        """Adiciona botão de importação na lista de estudantes"""
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

    def import_students_view(self, request):
        """View para upload e processamento da planilha"""
        if request.method == 'POST':
            try:
                excel_file = request.FILES.get('excel_file')

                if not excel_file:
                    messages.error(request, 'Por favor, selecione um arquivo.')
                    return redirect('..')

                # Valida extensão do arquivo
                if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
                    messages.error(request, 'Formato de arquivo inválido. Use .xlsx, .xls ou .csv')
                    return redirect('..')

                # Lê o arquivo
                if excel_file.name.endswith('.csv'):
                    df = pd.read_csv(excel_file)
                else:
                    df = pd.read_excel(excel_file)

                # Valida colunas obrigatórias
                required_columns = ['student_serial', 'student_name', 'class_name']
                missing_columns = [col for col in required_columns if col not in df.columns]

                if missing_columns:
                    messages.error(
                        request,
                        f'Colunas obrigatórias ausentes: {", ".join(missing_columns)}'
                    )
                    return redirect('..')

                # Processa as linhas
                success_count = 0
                error_count = 0
                errors = []

                for index, row in df.iterrows():
                    try:
                        # Busca a turma pelo nome
                        try:
                            class_obj = TbClass.objects.get(class_name=row['class_name'])
                        except TbClass.DoesNotExist:
                            errors.append(f'Linha {index + 2}: Turma "{row["class_name"]}" não encontrada')
                            error_count += 1
                            continue

                        # Prepara os dados do estudante
                        student_data = {
                            'student_serial': int(row['student_serial']),
                            'student_name': str(row['student_name']).strip(),
                            'id_class': class_obj,
                            'status': row.get('status', 'enrolled'),
                        }

                        # Adiciona data de matrícula se fornecida
                        if 'enrollment_date' in row and pd.notna(row['enrollment_date']):
                            try:
                                if isinstance(row['enrollment_date'], str):
                                    student_data['enrollment_date'] = datetime.strptime(
                                        row['enrollment_date'], '%Y-%m-%d'
                                    ).date()
                                else:
                                    student_data['enrollment_date'] = row['enrollment_date']
                            except ValueError:
                                errors.append(
                                    f'Linha {index + 2}: Data de matrícula inválida. Use formato AAAA-MM-DD'
                                )
                                error_count += 1
                                continue

                        # Verifica se o student_serial já existe
                        existing_student = TbStudents.objects.filter(
                            student_serial=student_data['student_serial']
                        ).first()

                        if existing_student:
                            # Atualiza estudante existente
                            for key, value in student_data.items():
                                if key != 'student_serial':
                                    setattr(existing_student, key, value)
                            existing_student.save()
                        else:
                            # Cria novo estudante
                            TbStudents.objects.create(**student_data)

                        success_count += 1

                    except Exception as e:
                        errors.append(f'Linha {index + 2}: {str(e)}')
                        error_count += 1

                # Mensagens de resultado
                if success_count > 0:
                    messages.success(
                        request,
                        f'{success_count} estudante(s) importado(s) com sucesso!'
                    )

                if error_count > 0:
                    error_message = f'{error_count} erro(s) encontrado(s):<br>'
                    error_message += '<br>'.join(errors[:10])  # Mostra apenas os primeiros 10 erros
                    if len(errors) > 10:
                        error_message += f'<br>... e mais {len(errors) - 10} erro(s)'
                    messages.error(request, format_html(error_message))

                return redirect('..')

            except Exception as e:
                messages.error(request, f'Erro ao processar arquivo: {str(e)}')
                return redirect('..')

        # GET request - mostra o formulário
        context = {
            'site_header': 'Importação de Estudantes',
            'title': 'Importar Estudantes em Lote',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }

        return render(request, 'admin/students/import_students.html', context)

    def download_template(self, request):
        """Gera e faz download de uma planilha modelo"""
        # Cria um workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudantes"

        # Cabeçalhos
        headers = ['student_serial', 'student_name', 'class_name', 'enrollment_date', 'status']
        ws.append(headers)

        # Exemplos de dados
        ws.append([12345, 'João da Silva', 'Turma A - 5º Ano', '2025-01-15', 'enrolled'])
        ws.append([12346, 'Maria Santos', 'Turma B - 6º Ano', '2025-01-15', 'enrolled'])
        ws.append([12347, 'Pedro Oliveira', 'Turma A - 5º Ano', '2025-01-16', 'enrolled'])

        # Formata os cabeçalhos
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

        # Adiciona uma aba de instruções
        ws_instructions = wb.create_sheet("Instruções")
        instructions = [
            ["INSTRUÇÕES PARA IMPORTAÇÃO DE ESTUDANTES"],
            [""],
            ["Colunas obrigatórias:"],
            ["- student_serial: Número de matrícula único do estudante (número inteiro)"],
            ["- student_name: Nome completo do estudante"],
            ["- class_name: Nome exato da turma (deve existir no sistema)"],
            [""],
            ["Colunas opcionais:"],
            ["- enrollment_date: Data de matrícula no formato AAAA-MM-DD (ex: 2025-01-15)"],
            ["- status: Status do estudante (padrão: enrolled)"],
            [""],
            ["Observações:"],
            ["- O student_serial deve ser único. Se já existir, o estudante será atualizado"],
            ["- O class_name deve corresponder exatamente a uma turma existente no sistema"],
            ["- Não altere os nomes das colunas na planilha"],
            ["- Você pode adicionar quantas linhas forem necessárias"],
        ]

        for row in instructions:
            ws_instructions.append(row)

        # Formata as instruções
        ws_instructions.column_dimensions['A'].width = 80

        # Salva em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Retorna o arquivo
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_estudantes.xlsx'

        return response


# Registra outros modelos
@admin.register(TbClass)
class TbClassAdmin(admin.ModelAdmin):
    list_display = ['id', 'class_name', 'id_school', 'id_teacher', 'grade', 'school_year', 'shift']
    list_filter = ['school_year', 'grade', 'shift', 'id_school']
    search_fields = ['class_name']


@admin.register(TbSchool)
class TbSchoolAdmin(admin.ModelAdmin):
    list_display = ['id', 'school', 'director_name', 'id_city', 'address']
    list_filter = ['id_city']
    search_fields = ['school', 'director_name']


@admin.register(TbCity)
class TbCityAdmin(admin.ModelAdmin):
    list_display = ['id', 'city', 'state']
    list_filter = ['state']
    search_fields = ['city']


@admin.register(TbTeacher)
class TbTeacherAdmin(admin.ModelAdmin):
    list_display = ['id', 'teacher_serial', 'teacher_name', 'status']
    list_filter = ['status']
    search_fields = ['teacher_name', 'teacher_serial']


@admin.register(TbSubject)
class TbSubjectAdmin(admin.ModelAdmin):
    list_display = ['id', 'subject_name', 'description']
    search_fields = ['subject_name']


@admin.register(TbTeacherSubject)
class TbTeacherSubjectAdmin(admin.ModelAdmin):
    list_display = ['id', 'id_teacher', 'id_subject']
    list_filter = ['id_teacher', 'id_subject']


@admin.register(TbTeacherSchool)
class TbTeacherSchoolAdmin(admin.ModelAdmin):
    list_display = ['id', 'id_teacher', 'id_school', 'assignment_date', 'status']
    list_filter = ['status', 'id_school']
    search_fields = ['id_teacher__teacher_name']
